import os
import shutil
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle
import pandas as pd
import re

# Source directory where Excel files are located
source_dir = 'C:\\Users\\jawad.khan\\Desktop'

# Function to create folders and process Excel files
def process_files():
    # Get list of Excel files in the source directory
    excel_files = [f for f in os.listdir(source_dir) if f.endswith('.xlsx')]

    # Regular expression to extract the date range from the filename
    date_pattern = r'\d{8} - \d{8}'

    # Dictionary to group files by their date range
    date_groups = {}

    for file in excel_files:
        match = re.search(date_pattern, file)
        if match:
            date_range = match.group(0)
            if date_range not in date_groups:
                date_groups[date_range] = []
            date_groups[date_range].append(file)

    # Process each group of files
    for date_range, files in date_groups.items():
        # Create a folder named after the date range
        folder_path = os.path.join(source_dir, date_range)
        os.makedirs(folder_path, exist_ok=True)

        # Move the relevant files to the folder
        for file in files:
            shutil.move(os.path.join(source_dir, file), os.path.join(folder_path, file))

        # Create a consolidated Excel report for the date range
        consolidated_report = os.path.join(folder_path, f'{date_range} OMI Report.xlsx')

        with pd.ExcelWriter(consolidated_report, engine='openpyxl') as writer:
            for file in files:
                file_path = os.path.join(folder_path, file)
                try:
                    # Open each Excel file
                    wb = load_workbook(file_path)
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]

                        # Create a dataframe to write to the consolidated report
                        data = sheet.values
                        cols = next(data)
                        df = pd.DataFrame(data, columns=cols)

                        # Write the dataframe to a new sheet with the original file's name
                        df.to_excel(writer, sheet_name=os.path.splitext(file)[0], index=False)
                    
                except InvalidFileException as e:
                    print(f"Error processing file {file}: {e}")

    print("Process completed.")

# Call the function to process the files
process_files()
